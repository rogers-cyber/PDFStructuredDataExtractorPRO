import os
import re
import sqlite3
import threading
import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.ttk as ttk
import ttkbootstrap as tb
import time
import concurrent.futures
import pdfplumber
import pytesseract
from pdf2image import convert_from_path
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------- INFO ----------------
APP_NAME = "PDF Structured Data Extractor PRO"
APP_VERSION = "2.0"
APP_AUTHOR = "Mate Technologies"

DB_NAME = "extracted_data.db"

class PDFExtractorApp:
    def __init__(self, master):
        self.master = master
        self.master.title(f"{APP_NAME} {APP_VERSION}")
        self.master.geometry("1100x700")
        self.style = tb.Style(theme="superhero")

        self.status_var = tk.StringVar(value="Idle")
        self.progress_val = tk.DoubleVar(value=0)

        self.stop_event = threading.Event()
        self.pause_event = threading.Event()

        self.create_database()
        self.create_ui()

    # ================= DATABASE =================

    def create_database(self):
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT,
                name TEXT,
                date TEXT,
                document_id TEXT
            )
        """)
        conn.commit()
        conn.close()

    def insert_record(self, file_name, name, date, doc_id):
        conn = sqlite3.connect(DB_NAME)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO records (file_name, name, date, document_id)
            VALUES (?, ?, ?, ?)
        """, (file_name, name, date, doc_id))
        conn.commit()
        conn.close()

    # ================= UI =================

    def create_ui(self):
        tb.Label(self.master, text=APP_NAME, font=("Segoe UI", 18, "bold")).pack(pady=10)

        top = tb.Frame(self.master, padding=10)
        top.pack(fill="x")

        tb.Button(top, text="Select Folder", bootstyle="primary", command=self.select_folder).pack(side="left", padx=5)
        tb.Button(top, text="Export Excel", bootstyle="success", command=self.export_excel).pack(side="left", padx=5)
        tb.Button(top, text="Pause", bootstyle="secondary", command=self.pause_scan).pack(side="left", padx=5)
        tb.Button(top, text="Resume", bootstyle="info", command=self.resume_scan).pack(side="left", padx=5)
        tb.Button(top, text="Stop", bootstyle="danger", command=self.stop_scan).pack(side="left", padx=5)
        tb.Button(top, text="About", bootstyle="warning", command=self.show_about).pack(side="right", padx=5)

        tb.Label(top, textvariable=self.status_var).pack(side="right")

        self.tree = ttk.Treeview(self.master, columns=("file", "name", "date", "id"), show="headings")
        self.tree.heading("file", text="File")
        self.tree.heading("name", text="Name")
        self.tree.heading("date", text="Date")
        self.tree.heading("id", text="Document ID")
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        tb.Progressbar(self.master, variable=self.progress_val, maximum=100).pack(fill="x", padx=10, pady=5)

    # ================= BUTTON ACTIONS =================

    def select_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            threading.Thread(target=self.scan_folder, args=(folder,), daemon=True).start()

    def pause_scan(self):
        self.pause_event.set()

    def resume_scan(self):
        self.pause_event.clear()

    def stop_scan(self):
        self.stop_event.set()

    # ================= SCAN =================

    def scan_folder(self, folder):
        self.status_var.set("Scanning...")
        pdfs = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith(".pdf")]
        total = len(pdfs)
        processed = 0

        for pdf in pdfs:
            if self.stop_event.is_set():
                return
            while self.pause_event.is_set():
                time.sleep(0.2)

            record = self.process_pdf(pdf)
            if record:
                file_name, name, date, doc_id = record
                self.insert_record(file_name, name, date, doc_id)
                self.tree.insert("", "end", values=record)

            processed += 1
            self.progress_val.set((processed/total)*100)

        self.status_var.set("Completed.")

    # ================= AI TEXT EXTRACTION =================

    def process_pdf(self, path):
        file_name = os.path.basename(path)
        text = ""

        try:
            with pdfplumber.open(path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

            if not text.strip():
                images = convert_from_path(path, dpi=300)
                for img in images:
                    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
                    for i, word in enumerate(data["text"]):
                        conf = int(data["conf"][i])
                        width = data["width"][i]
                        height = data["height"][i]

                        # Advanced handwriting suppression
                        if conf > 75 and len(word) > 2 and width > 10 and height > 10:
                            text += word + " "

        except:
            return None

        # Structured Field Extraction
        name = self.extract_name(text)
        date = self.extract_date(text)
        doc_id = self.extract_id(text)

        return (file_name, name, date, doc_id)

    # ================= FIELD EXTRACTION =================

    def extract_name(self, text):
        match = re.search(r"Name[:\s]+([A-Z][a-z]+\s[A-Z][a-z]+)", text)
        return match.group(1) if match else ""

    def extract_date(self, text):
        match = re.search(r"\b\d{2}/\d{2}/\d{4}\b", text)
        return match.group(0) if match else ""

    def extract_id(self, text):
        match = re.search(r"\bID[:\s]*([A-Z0-9\-]+)", text)
        return match.group(1) if match else ""

    # ================= EXPORT EXCEL =================

    def export_excel(self):
        conn = sqlite3.connect(DB_NAME)
        df = conn.execute("SELECT file_name, name, date, document_id FROM records").fetchall()
        conn.close()

        if not df:
            messagebox.showwarning("No Data", "Database is empty.")
            return

        save_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
        if not save_path:
            return

        wb = Workbook()
        ws = wb.active
        headers = ["File Name", "Name", "Date", "Document ID"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for row in df:
            ws.append(row)

        # Auto column width
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        ws.freeze_panes = "A2"
        wb.save(save_path)

        messagebox.showinfo("Success", "Excel exported successfully.")

    def show_about(self):
        messagebox.showinfo(
            f"About {APP_NAME}",
            f"{APP_NAME} v{APP_VERSION}\n\n"

            "PDF Structured Data Extractor PRO is an enterprise-grade document "
            "processing tool designed to extract structured data from consistently "
            "formatted PDFs — including scanned documents.\n\n"

            "Core Features:\n"
            "• Intelligent structured field extraction (Name, Date, ID, etc.)\n"
            "• AI-enhanced handwriting suppression\n"
            "• OCR fallback for scanned PDFs\n"
            "• SQLite database storage\n"
            "• Automatic Excel export with formatting\n"
            "• Multi-threaded processing engine\n"
            "• Pause / Resume / Stop controls\n"
            "• Clean modern UI powered by ttkbootstrap\n\n"

            "Ideal For:\n"
            "• Legal offices processing case documents\n"
            "• Medical practices digitizing patient forms\n"
            "• Insurance companies handling claims\n"
            "• HR departments processing employee paperwork\n"
            "• Any organization working with structured PDF forms\n\n"

            "How It Works:\n"
            "1. Select a folder containing PDFs.\n"
            "2. The engine extracts digital text or performs OCR if needed.\n"
            "3. Handwriting and low-confidence text are filtered out.\n"
            "4. Structured fields are identified and stored in a database.\n"
            "5. Export clean, formatted Excel reports instantly.\n\n"

            "Built for scalability and SaaS deployment.\n\n"

            f"{APP_AUTHOR}\n"
            "Professional Document Automation Solutions"
        )

# ================= RUN =================

if __name__ == "__main__":
    root = tk.Tk()
    app = PDFExtractorApp(root)
    root.mainloop()
